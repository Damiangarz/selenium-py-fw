import openpyxl


class SearchItemTestData:
    search_item_data_set = [{'search_item': 'ps4 consola'},
                            {'search_item': 'x box consola'},
                            {'search_item': 'nintendo switch consola'}]
    search_item_data = [{'search_item': 'ps4 consola'}]

    @staticmethod
    def get_excel_data(test_case_name):
        book = openpyxl.load_workbook("C:\\Frameworks\\pythonSelFW1\\data_sheet.xlsx")
        sheet = book.active
        data_dict = {}
        for i in range(1, sheet.max_row + 1):
            if sheet.cell(row=i, column=1).value == test_case_name:
                for j in range(2, sheet.max_column + 1):
                    data_dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
        return [data_dict]
